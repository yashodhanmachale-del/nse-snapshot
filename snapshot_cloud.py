"""
snapshot_cloud.py v10
Angel One Smart API → Excel (Nifty 50 snapshot) + Breadth Score + Email + Drive + Sheets
"""

import os, json, pyotp, gspread, traceback, smtplib, time, requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from SmartApi import SmartConnect
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

IST = ZoneInfo("Asia/Kolkata")
INSTRUMENT_MASTER_URL = "https://margincalculator.angelone.in/OpenAPI_File/files/OpenAPIScripMaster.json"

# ════════════════════════════════════════════════════════════
# BREADTH SCORE FORMULA: count * 7 / 50  → round to int
# Positive score + Negative score = 7 always
# ════════════════════════════════════════════════════════════
def breadth_score(count):
    return round(count * 7 / 50)

def calc_breadth(stocks):
    """
    Breadth score formula: count * 7 / 50 → round to integer.
    Ranked display: higher score always listed first.
    ranked_stocks: top N stocks from dominant side + top M from other side = always 7 total.
    """
    valid   = [s for s in stocks if s.get("contrib") is not None]
    pos     = [s for s in valid if s["contrib"] >= 0]
    neg     = [s for s in valid if s["contrib"] <  0]
    pos_cnt = len(pos)
    neg_cnt = len(neg)
    pos_pts = round(sum(s["contrib"] for s in pos), 2)
    neg_pts = round(sum(s["contrib"] for s in neg), 2)
    pos_sc  = breadth_score(pos_cnt)
    neg_sc  = breadth_score(neg_cnt)

    pos_stocks_sorted = sorted(pos, key=lambda x: x["contrib"], reverse=True)
    neg_stocks_sorted = sorted(neg, key=lambda x: x["contrib"])

    # Ranked: higher score first
    if pos_sc >= neg_sc:
        dominant      = "POSITIVE"
        first_score   = pos_sc;   second_score  = neg_sc
        first_label   = "🟢 POSITIVE"; second_label = "🔴 NEGATIVE"
        first_color   = "1B5E20";  second_color  = "B71C1C"
        first_bg      = "E8F5E9";  second_bg     = "FFEBEE"
        first_count   = pos_cnt;   second_count  = neg_cnt
        first_pts     = pos_pts;   second_pts    = neg_pts
        ranked_stocks = (
            [{"rank":i+1,"symbol":s["symbol"],"contrib":s["contrib"],
              "pChng":s["pChng"],"ltp":s["ltp"],"side":"POSITIVE",
              "color":"1B5E20","bg":"E8F5E9"}
             for i,s in enumerate(pos_stocks_sorted[:pos_sc])]
            +
            [{"rank":pos_sc+i+1,"symbol":s["symbol"],"contrib":s["contrib"],
              "pChng":s["pChng"],"ltp":s["ltp"],"side":"NEGATIVE",
              "color":"B71C1C","bg":"FFEBEE"}
             for i,s in enumerate(neg_stocks_sorted[:neg_sc])]
        )
    else:
        dominant      = "NEGATIVE"
        first_score   = neg_sc;   second_score  = pos_sc
        first_label   = "🔴 NEGATIVE"; second_label = "🟢 POSITIVE"
        first_color   = "B71C1C";  second_color  = "1B5E20"
        first_bg      = "FFEBEE";  second_bg     = "E8F5E9"
        first_count   = neg_cnt;   second_count  = pos_cnt
        first_pts     = neg_pts;   second_pts    = pos_pts
        ranked_stocks = (
            [{"rank":i+1,"symbol":s["symbol"],"contrib":s["contrib"],
              "pChng":s["pChng"],"ltp":s["ltp"],"side":"NEGATIVE",
              "color":"B71C1C","bg":"FFEBEE"}
             for i,s in enumerate(neg_stocks_sorted[:neg_sc])]
            +
            [{"rank":neg_sc+i+1,"symbol":s["symbol"],"contrib":s["contrib"],
              "pChng":s["pChng"],"ltp":s["ltp"],"side":"POSITIVE",
              "color":"1B5E20","bg":"E8F5E9"}
             for i,s in enumerate(pos_stocks_sorted[:pos_sc])]
        )

    return {
        "pos_count":       pos_cnt,
        "neg_count":       neg_cnt,
        "pos_contrib_sum": pos_pts,
        "neg_contrib_sum": neg_pts,
        "pos_score":       pos_sc,
        "neg_score":       neg_sc,
        "total_score":     pos_sc + neg_sc,
        "display":         f"{pos_sc} | {neg_sc}",
        "pos_stocks":      pos_stocks_sorted,
        "neg_stocks":      neg_stocks_sorted,
        "ranked_stocks":   ranked_stocks,
        "dominant":        dominant,
        "first_score":     first_score,
        "second_score":    second_score,
        "first_label":     first_label,
        "second_label":    second_label,
        "first_color":     first_color,
        "second_color":    second_color,
        "first_bg":        first_bg,
        "second_bg":       second_bg,
        "first_count":     first_count,
        "second_count":    second_count,
        "first_pts":       first_pts,
        "second_pts":      second_pts,
        "ranked_display":  f"{first_score} | {second_score}",
    }

# ════════════════════════════════════════════════════════════
# CREDENTIALS
# ════════════════════════════════════════════════════════════
def env(key, default=""):
    return os.environ.get(key, default).strip()

def get_creds():
    return {
        "api_key":      env("ANGEL_API_KEY"),
        "client_id":    env("ANGEL_CLIENT_ID"),
        "password":     env("ANGEL_PASSWORD"),
        "totp_secret":  env("ANGEL_TOTP_SECRET"),
        "gmail_sender": env("GMAIL_SENDER"),
        "gmail_pass":   env("GMAIL_APP_PASSWORD"),
        "recipient":    env("RECIPIENT_EMAIL"),
        "drive_folder": env("GOOGLE_DRIVE_FOLDER_ID"),
        "sheet_id":     env("GOOGLE_SHEET_ID"),
        "sa_json":      env("SERVICE_ACCOUNT_JSON"),

    }

def get_sa_creds(sa_json_str, scopes):
    return Credentials.from_service_account_info(json.loads(sa_json_str), scopes=scopes)

# ════════════════════════════════════════════════════════════
# EXCEL STYLES
# ════════════════════════════════════════════════════════════
C = {
    "title_bg":"0D47A1","hdr_mid":"1976D2","hdr_dark":"283593",
    "pos_bg":"E8F5E9","pos_alt":"F1F8E9","pos_txt":"1B5E20",
    "neg_bg":"FFEBEE","neg_alt":"FCE4EC","neg_txt":"B71C1C",
    "idx_bg":"E8EAF6","idx_alt":"FFFFFF",
    "grn_hdr":"2E7D32","red_hdr":"C62828","white":"FFFFFF",
    "yellow":"FFF176","yellow_lt":"FFF9C4",
}
def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
def font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")
def aln(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def brd():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)
def sc(cell, bg=None, fg="000000", bold=False, size=10, ha="left"):
    if bg: cell.fill = fill(bg)
    cell.font = font(bold, fg, size)
    cell.alignment = aln(ha)
    cell.border = brd()

# ════════════════════════════════════════════════════════════
# ANGEL ONE
# ════════════════════════════════════════════════════════════
def angel_login(creds):
    print("🔑 Logging in to Angel One...")
    obj  = SmartConnect(api_key=creds["api_key"])
    totp = pyotp.TOTP(creds["totp_secret"]).now()
    data = obj.generateSession(creds["client_id"], creds["password"], totp)
    if data["status"] is False:
        raise Exception("Login failed: " + str(data.get("message","")))
    print("✅ Login OK")
    return obj

# Hardcoded fallback tokens for all 50 stocks (used if instrument master download fails)
FALLBACK_TOKENS = {
    "RELIANCE":{"token":"2885","tradingsymbol":"RELIANCE-EQ"},
    "HDFCBANK":{"token":"1333","tradingsymbol":"HDFCBANK-EQ"},
    "ICICIBANK":{"token":"4963","tradingsymbol":"ICICIBANK-EQ"},
    "INFY":{"token":"1594","tradingsymbol":"INFY-EQ"},
    "TCS":{"token":"11536","tradingsymbol":"TCS-EQ"},
    "BHARTIARTL":{"token":"10604","tradingsymbol":"BHARTIARTL-EQ"},
    "ITC":{"token":"1660","tradingsymbol":"ITC-EQ"},
    "LT":{"token":"11483","tradingsymbol":"LT-EQ"},
    "KOTAKBANK":{"token":"1922","tradingsymbol":"KOTAKBANK-EQ"},
    "AXISBANK":{"token":"5900","tradingsymbol":"AXISBANK-EQ"},
    "SBIN":{"token":"3045","tradingsymbol":"SBIN-EQ"},
    "HINDUNILVR":{"token":"1394","tradingsymbol":"HINDUNILVR-EQ"},
    "BAJFINANCE":{"token":"317","tradingsymbol":"BAJFINANCE-EQ"},
    "M&M":{"token":"2031","tradingsymbol":"M&M-EQ"},
    "MARUTI":{"token":"10999","tradingsymbol":"MARUTI-EQ"},
    "SUNPHARMA":{"token":"3351","tradingsymbol":"SUNPHARMA-EQ"},
    "ULTRACEMCO":{"token":"11532","tradingsymbol":"ULTRACEMCO-EQ"},
    "HCLTECH":{"token":"7229","tradingsymbol":"HCLTECH-EQ"},
    "TITAN":{"token":"3506","tradingsymbol":"TITAN-EQ"},
    "BAJAJFINSV":{"token":"16675","tradingsymbol":"BAJAJFINSV-EQ"},
    "ASIANPAINT":{"token":"236","tradingsymbol":"ASIANPAINT-EQ"},
    "ADANIENT":{"token":"25","tradingsymbol":"ADANIENT-EQ"},
    "NTPC":{"token":"11630","tradingsymbol":"NTPC-EQ"},
    "ONGC":{"token":"2475","tradingsymbol":"ONGC-EQ"},
    "POWERGRID":{"token":"14977","tradingsymbol":"POWERGRID-EQ"},
    "TATASTEEL":{"token":"3499","tradingsymbol":"TATASTEEL-EQ"},
    "TECHM":{"token":"13538","tradingsymbol":"TECHM-EQ"},
    "WIPRO":{"token":"3787","tradingsymbol":"WIPRO-EQ"},
    "ADANIPORTS":{"token":"15083","tradingsymbol":"ADANIPORTS-EQ"},
    "COALINDIA":{"token":"20374","tradingsymbol":"COALINDIA-EQ"},
    "JSWSTEEL":{"token":"11723","tradingsymbol":"JSWSTEEL-EQ"},
    "BAJAJ-AUTO":{"token":"16669","tradingsymbol":"BAJAJ-AUTO-EQ"},
    "GRASIM":{"token":"1232","tradingsymbol":"GRASIM-EQ"},
    "DRREDDY":{"token":"881","tradingsymbol":"DRREDDY-EQ"},
    "CIPLA":{"token":"694","tradingsymbol":"CIPLA-EQ"},
    "EICHERMOT":{"token":"910","tradingsymbol":"EICHERMOT-EQ"},
    "INDUSINDBK":{"token":"5258","tradingsymbol":"INDUSINDBK-EQ"},
    "APOLLOHOSP":{"token":"157","tradingsymbol":"APOLLOHOSP-EQ"},
    "TATACONSUM":{"token":"3432","tradingsymbol":"TATACONSUM-EQ"},
    "BEL":{"token":"383","tradingsymbol":"BEL-EQ"},
    "SHRIRAMFIN":{"token":"4306","tradingsymbol":"SHRIRAMFIN-EQ"},
    "HINDALCO":{"token":"1363","tradingsymbol":"HINDALCO-EQ"},
    "HDFCLIFE":{"token":"467","tradingsymbol":"HDFCLIFE-EQ"},
    "INDIGO":{"token":"11195","tradingsymbol":"INDIGO-EQ"},
    "JIOFIN":{"token":"543257","tradingsymbol":"JIOFIN-EQ"},
    "MAXHEALTH":{"token":"27466","tradingsymbol":"MAXHEALTH-EQ"},
    "SBILIFE":{"token":"21808","tradingsymbol":"SBILIFE-EQ"},
    "TMPV":{"token":"3456","tradingsymbol":"TATAMOTORS-EQ"},
    "TRENT":{"token":"1964","tradingsymbol":"TRENT-EQ"},
    "ETERNAL":{"token":"5097","tradingsymbol":"ZOMATO-EQ"},
}

def load_token_map():
    print("\n📥 Downloading instrument master...")
    for attempt in range(3):
        try:
            r = requests.get(INSTRUMENT_MASTER_URL, timeout=30)
            r.raise_for_status()
            data = r.json()
            token_map = {}
            for item in data:
                if item.get("exch_seg") == "NSE" and item.get("symbol","").endswith("-EQ"):
                    base = item["symbol"].replace("-EQ","")
                    token_map[base] = {"token": item["token"], "tradingsymbol": item["symbol"]}
            print(f"  ✅ {len(token_map)} NSE equities loaded from instrument master")
            return token_map
        except Exception as e:
            print(f"  ⚠️  Attempt {attempt+1} failed: {e}")
            if attempt < 2:
                time.sleep(3)
    print("  ⚠️  Instrument master unavailable — using hardcoded fallback tokens")
    return FALLBACK_TOKENS

INDEX_TOKENS = {
    "NIFTY 50":       ("NSE", "Nifty 50",          "99926000"),
    "SENSEX":         ("BSE", "SENSEX",             "99919000"),
    "BANK NIFTY":     ("NSE", "Nifty Bank",         "99926009"),
    "NIFTY IT":       ("NSE", "Nifty IT",           "99926011"),
    "NIFTY SMALLCAP 50": ("NSE", "Nifty Smallcap 50", "99926061"),
}

def fetch_one_index(obj, name, exch, sym, token):
    try:
        r = obj.ltpData(exch, sym, token)
        if r.get("status") and r.get("data"):
            ltp   = float(r["data"].get("ltp",   0) or 0)
            close = float(r["data"].get("close", ltp) or ltp)
            chng  = round(ltp - close, 2)
            pct   = round((chng / close) * 100, 2) if close else 0.0
            return name, {"ltp": ltp, "chng": chng, "pct": pct, "close": close}
        return name, {"ltp": None, "chng": None, "pct": None, "close": None}
    except Exception as e:
        print(f"  ❌ {name}: {e}")
        return name, {"ltp": None, "chng": None, "pct": None, "close": None}

def fetch_indices(obj):
    print("\n📊 Fetching indices in parallel...")
    result = {}
    with ThreadPoolExecutor(max_workers=5) as ex:
        futures = {ex.submit(fetch_one_index, obj, name, exch, sym, token): name
                   for name, (exch, sym, token) in INDEX_TOKENS.items()}
        for future in as_completed(futures):
            name, data = future.result()
            result[name] = data
            pct = data.get("pct")
            ltp = data.get("ltp")
            if ltp:
                print(f"  ✅ {name}: {ltp}  {pct:+.2f}%")
            else:
                print(f"  ⚠️  {name}: no data")
    return result

NIFTY50_LIST = {
    "RELIANCE":9.25,"HDFCBANK":12.80,"ICICIBANK":8.85,"INFY":5.85,"TCS":3.90,
    "BHARTIARTL":4.20,"ITC":3.55,"LT":3.45,"KOTAKBANK":2.95,"AXISBANK":2.85,
    "SBIN":2.75,"HINDUNILVR":2.30,"BAJFINANCE":2.05,"M&M":1.95,"MARUTI":1.55,
    "SUNPHARMA":1.50,"ULTRACEMCO":1.35,"HCLTECH":1.30,"TITAN":1.25,"BAJAJFINSV":1.20,
    "ASIANPAINT":1.15,"ADANIENT":1.10,"NTPC":1.10,"ONGC":1.05,"POWERGRID":1.05,
    "TATASTEEL":1.00,"TECHM":0.95,"WIPRO":0.90,"ADANIPORTS":0.90,"COALINDIA":0.85,
    "JSWSTEEL":0.80,"BAJAJ-AUTO":0.80,"GRASIM":0.70,"DRREDDY":0.65,"CIPLA":0.65,
    "EICHERMOT":0.65,"INDUSINDBK":0.60,"APOLLOHOSP":0.55,"TATACONSUM":0.55,
    "BEL":0.50,"SHRIRAMFIN":0.45,"HINDALCO":0.75,"HDFCLIFE":0.90,"INDIGO":0.70,
    "JIOFIN":0.85,"MAXHEALTH":0.55,"SBILIFE":0.85,"TMPV":0.60,"TRENT":0.95,
    "ETERNAL":0.95,
}

# ════════════════════════════════════════════════════════════
# SPECIAL WATCHLIST — Always tracked separately
# ════════════════════════════════════════════════════════════
SPECIAL_WATCHLIST = [
    "RELIANCE", "MARUTI", "HDFCBANK", "BHARTIARTL", "INFY",
    "TCS", "HINDUNILVR", "SUNPHARMA", "M&M", "ICICIBANK", "SBIN"
]

def get_special_stocks(stocks):
    """Extract special watchlist stocks from full stocks list."""
    stock_map = {s["symbol"]: s for s in stocks}
    return [stock_map[sym] for sym in SPECIAL_WATCHLIST if sym in stock_map]


def fetch_one_stock(obj, sym, wt, info, nifty_ltp):
    """Fetch a single stock — called in parallel."""
    if not info:
        wpts = round(nifty_ltp*(wt/100),2) if nifty_ltp else None
        return {"symbol":sym,"ltp":None,"chng":None,"pChng":None,
                "weight":wt,"weight_pts":wpts,"contrib":None}
    for attempt in range(3):
        try:
            r = obj.ltpData("NSE", info["tradingsymbol"], info["token"])
            if r.get("status") and r.get("data"):
                ltp   = float(r["data"].get("ltp",   0) or 0)
                close = float(r["data"].get("close", ltp) or ltp)
                chng  = round(ltp - close, 2)
                pct   = round((chng / close) * 100, 2) if close else 0.0
                contrib   = round(nifty_ltp * (pct/100) * (wt/100), 2) if nifty_ltp else None
                weight_pts= round(nifty_ltp * (wt/100), 2) if nifty_ltp else None
                return {"symbol":sym,"ltp":ltp,"chng":chng,"pChng":pct,
                        "weight":wt,"weight_pts":weight_pts,"contrib":contrib}
            else:
                msg = r.get("message","")
                if "rate" in msg.lower() and attempt < 2:
                    time.sleep(2); continue
                wpts = round(nifty_ltp*(wt/100),2) if nifty_ltp else None
                return {"symbol":sym,"ltp":None,"chng":None,"pChng":None,
                        "weight":wt,"weight_pts":wpts,"contrib":None}
        except Exception as e:
            if "rate" in str(e).lower() and attempt < 2:
                time.sleep(3); continue
            wpts = round(nifty_ltp*(wt/100),2) if nifty_ltp else None
            return {"symbol":sym,"ltp":None,"chng":None,"pChng":None,
                    "weight":wt,"weight_pts":wpts,"contrib":None}
    wpts = round(nifty_ltp*(wt/100),2) if nifty_ltp else None
    return {"symbol":sym,"ltp":None,"chng":None,"pChng":None,
            "weight":wt,"weight_pts":wpts,"contrib":None}

def fetch_nifty50(obj, nifty_ltp, token_map):
    print("\n📋 Fetching Nifty 50 stocks in parallel (10 at a time)...")
    symbols = list(NIFTY50_LIST.items())
    results = {}

    # Fetch in batches of 10 parallel — avoids rate limit
    BATCH = 10
    for i in range(0, len(symbols), BATCH):
        batch = symbols[i:i+BATCH]
        with ThreadPoolExecutor(max_workers=BATCH) as ex:
            futures = {ex.submit(fetch_one_stock, obj, sym, wt,
                                 token_map.get(sym), nifty_ltp): sym
                       for sym, wt in batch}
            for future in as_completed(futures):
                s = future.result()
                results[s["symbol"]] = s
                if s["ltp"]:
                    print(f"  ✅ {s['symbol']}: {s['ltp']}  {s['pChng']:+.2f}%")
        # Small pause between batches to avoid rate limit
        if i + BATCH < len(symbols):
            time.sleep(1)

    # Return in original NIFTY50_LIST order
    stocks = [results[sym] for sym, _ in symbols if sym in results]
    filled = sum(1 for s in stocks if s["ltp"] is not None)
    print(f"\n  Result: {filled}/{len(stocks)} stocks with data")
    return stocks

# ════════════════════════════════════════════════════════════
# LEVEL CALCULATOR
# ════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════
# BREADTH SCORE
# Formula: count * 7 / 50 → round to integer
# Positive score + Negative score always = 7
# ════════════════════════════════════════════════════════════
def build_snapshot_excel(label, ist_dt, indices, stocks, breadth, special_stocks):
    wb = Workbook(); ws = wb.active
    ws.title = "Snapshot_" + label
    disp = label[:2]+":"+label[2:] if len(label)==4 else label
    time_str = ist_dt.strftime("%d-%b-%Y %H:%M:%S")

    ws.merge_cells("A1:P1")
    c = ws["A1"]; c.value = f"NSE Market Snapshot  —  {disp} IST"
    sc(c, bg=C["title_bg"], fg=C["white"], bold=True, size=14, ha="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:P2")
    ws["A2"].value = f"Captured At (IST):  {time_str}"
    ws["A2"].fill = fill("E3F2FD"); ws["A2"].font = font(size=10, color="333333")
    ws["A2"].alignment = aln("center")

    # ── BREADTH SCORE SUMMARY TABLE ─────────────────────────
    ws.merge_cells("A4:P4")
    c = ws["A4"]
    pos_s = breadth["pos_score"]; neg_s = breadth["neg_score"]
    tot_s = breadth["total_score"]
    pos_c = breadth["pos_count"]; neg_c = breadth["neg_count"]
    pos_p = breadth["pos_contrib_sum"]; neg_p = breadth["neg_contrib_sum"]
    c.value = (f"📊  BREADTH SCORE:   "
               f"🟢 POSITIVE  {pos_s}  ({pos_c} stocks, {pos_p:+.2f} pts)   "
               f"|   "
               f"🔴 NEGATIVE  {neg_s}  ({neg_c} stocks, {neg_p:+.2f} pts)   "
               f"=   TOTAL  {tot_s}  (always 7)   "
               f"│  Formula: count × 7 ÷ 50")
    c.fill = fill("0D47A1"); c.font = font(bold=True, color="FFFFFF", size=11)
    c.alignment = aln("center"); c.border = brd()
    ws.row_dimensions[4].height = 24

    # Detailed breadth score boxes in row 5
    # Box 1: Positive Score (large)
    ws.merge_cells("A5:D6")
    c = ws["A5"]
    c.value = f"🟢  {pos_s}"
    c.fill = fill("1B5E20"); c.font = font(bold=True, color="FFFFFF", size=28)
    c.alignment = aln("center"); c.border = brd()
    ws.row_dimensions[5].height = 30; ws.row_dimensions[6].height = 22

    ws.merge_cells("E5:H6")
    c = ws["E5"]
    c.value = f"Positive: {pos_c} stocks  |  Contrib: {pos_p:+.2f} pts  |  Score={pos_c}x7/50={pos_s}"
    c.fill = fill("E8F5E9"); c.font = font(bold=False, color="1B5E20", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = brd()

    # Box 2: Separator |
    ws.merge_cells("I5:I6")
    c = ws["I5"]
    c.value = "|"
    c.fill = fill("E8EAF6"); c.font = font(bold=True, color="0D47A1", size=24)
    c.alignment = aln("center"); c.border = brd()

    # Box 3: Negative Score (large)
    ws.merge_cells("J5:M6")
    c = ws["J5"]
    c.value = f"🔴  {neg_s}"
    c.fill = fill("B71C1C"); c.font = font(bold=True, color="FFFFFF", size=28)
    c.alignment = aln("center"); c.border = brd()

    ws.merge_cells("N5:P6")
    c = ws["N5"]
    c.value = f"Negative: {neg_c} stocks  |  Contrib: {neg_p:+.2f} pts  |  Score={neg_c}x7/50={neg_s}"
    c.fill = fill("FFEBEE"); c.font = font(bold=False, color="B71C1C", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = brd()

    # ── RANKED BREADTH TABLE in Excel ───────────────────────
    r = 8
    ws.merge_cells(f"A{r}:P{r}")
    sc(ws.cell(r,1,"📋  BREADTH RANKING TABLE  (Higher Score Listed First)"),
       bg="37474F", fg="FFFFFF", bold=True, size=11, ha="center")
    r += 1
    for h, col in zip(["RANK","SIDE","SCORE","STOCKS","CONTRIB SUM (pts)","FORMULA",
                        "","","","","","","","","",""], range(1,17)):
        if h:
            sc(ws.cell(r,col,h), bg="0D47A1", fg="FFFFFF", bold=True, ha="center")

    # Row 1: First ranked (higher score)
    r += 1
    ws.merge_cells(f"A{r}:A{r}")
    sc(ws.cell(r,1,"1st"), bg=breadth["first_bg"], fg=breadth["first_color"], bold=True, size=14, ha="center")
    sc(ws.cell(r,2,breadth["first_label"]), bg=breadth["first_bg"], fg=breadth["first_color"], bold=True, ha="center")
    ws.cell(r,3,breadth["first_score"]).font = font(bold=True, color=breadth["first_color"], size=20)
    ws.cell(r,3).fill=fill(breadth["first_bg"]); ws.cell(r,3).border=brd(); ws.cell(r,3).alignment=aln("center")
    ws.cell(r,4,breadth["first_count"]).font = font(bold=True, color=breadth["first_color"], size=12)
    ws.cell(r,4).fill=fill(breadth["first_bg"]); ws.cell(r,4).border=brd(); ws.cell(r,4).alignment=aln("center")
    ws.cell(r,5,breadth["first_pts"]).number_format="+#,##0.00;-#,##0.00"
    ws.cell(r,5).font=font(bold=True, color=breadth["first_color"])
    ws.cell(r,5).fill=fill(breadth["first_bg"]); ws.cell(r,5).border=brd(); ws.cell(r,5).alignment=aln("center")
    ws.cell(r,6,f'{breadth["first_count"]}×7÷50={breadth["first_score"]}')
    ws.cell(r,6).font=font(color="555555"); ws.cell(r,6).fill=fill(breadth["first_bg"])
    ws.cell(r,6).border=brd(); ws.cell(r,6).alignment=aln("center")
    ws.row_dimensions[r].height = 24

    # Row 2: Second ranked (lower score)
    r += 1
    sc(ws.cell(r,1,"2nd"), bg=breadth["second_bg"], fg=breadth["second_color"], bold=True, size=14, ha="center")
    sc(ws.cell(r,2,breadth["second_label"]), bg=breadth["second_bg"], fg=breadth["second_color"], bold=True, ha="center")
    ws.cell(r,3,breadth["second_score"]).font=font(bold=True, color=breadth["second_color"], size=20)
    ws.cell(r,3).fill=fill(breadth["second_bg"]); ws.cell(r,3).border=brd(); ws.cell(r,3).alignment=aln("center")
    ws.cell(r,4,breadth["second_count"]).font=font(bold=True, color=breadth["second_color"], size=12)
    ws.cell(r,4).fill=fill(breadth["second_bg"]); ws.cell(r,4).border=brd(); ws.cell(r,4).alignment=aln("center")
    ws.cell(r,5,breadth["second_pts"]).number_format="+#,##0.00;-#,##0.00"
    ws.cell(r,5).font=font(bold=True, color=breadth["second_color"])
    ws.cell(r,5).fill=fill(breadth["second_bg"]); ws.cell(r,5).border=brd(); ws.cell(r,5).alignment=aln("center")
    ws.cell(r,6,f'{breadth["second_count"]}×7÷50={breadth["second_score"]}')
    ws.cell(r,6).font=font(color="555555"); ws.cell(r,6).fill=fill(breadth["second_bg"])
    ws.cell(r,6).border=brd(); ws.cell(r,6).alignment=aln("center")
    ws.row_dimensions[r].height = 24

    # Total row
    r += 1
    sc(ws.cell(r,1,"TOTAL"), bg="E8EAF6", fg="0D47A1", bold=True, ha="center")
    sc(ws.cell(r,2,f'Dominant: {breadth["dominant"]}'), bg="E8EAF6", fg="0D47A1", bold=True, ha="center")
    ws.cell(r,3,breadth["total_score"]).font=font(bold=True, color="0D47A1", size=16)
    ws.cell(r,3).fill=fill("E8EAF6"); ws.cell(r,3).border=brd(); ws.cell(r,3).alignment=aln("center")
    ws.cell(r,4,breadth["pos_count"]+breadth["neg_count"]).font=font(bold=True, color="0D47A1")
    ws.cell(r,4).fill=fill("E8EAF6"); ws.cell(r,4).border=brd(); ws.cell(r,4).alignment=aln("center")
    total_pts = round(breadth["pos_contrib_sum"]+breadth["neg_contrib_sum"],2)
    ws.cell(r,5,total_pts).number_format="+#,##0.00;-#,##0.00"
    ws.cell(r,5).font=font(bold=True, color="0D47A1")
    ws.cell(r,5).fill=fill("E8EAF6"); ws.cell(r,5).border=brd(); ws.cell(r,5).alignment=aln("center")
    sc(ws.cell(r,6,f'Ranked: {breadth["ranked_display"]}'), bg="E8EAF6", fg="0D47A1", bold=True, ha="center")
    ws.row_dimensions[r].height = 20

    r += 2

    # ── 7-STOCK RANKED TABLE in Excel ────────────────────────
    total_sc = breadth["total_score"]
    ws.merge_cells(f"A{r}:G{r}")
    sc(ws.cell(r,1,
        f"🏆  TOP {total_sc} STOCKS BY BREADTH SCORE  "
        f"({breadth['first_score']} {breadth['first_label']} + "
        f"{breadth['second_score']} {breadth['second_label']})"),
       bg="37474F", fg="FFFFFF", bold=True, size=11, ha="center")
    r += 1
    for h, col in zip(["#","SYMBOL","SIDE","LTP (₹)","% CHANGE","CONTRIBUTION (pts)",""], range(1,8)):
        sc(ws.cell(r,col,h), bg="0D47A1", fg="FFFFFF", bold=True, ha="center")
    ws.row_dimensions[r].height = 18

    for s in breadth["ranked_stocks"]:
        r += 1
        bg = s["bg"]; tc = s["color"]
        side_lbl = "🟢 POSITIVE" if s["side"]=="POSITIVE" else "🔴 NEGATIVE"

        ws.cell(r,1,s["rank"]).font=font(bold=True,color=tc,size=12)
        ws.cell(r,1).fill=fill(bg);ws.cell(r,1).border=brd();ws.cell(r,1).alignment=aln("center")

        ws.cell(r,2,s["symbol"]).font=font(bold=True,color=tc,size=13)
        ws.cell(r,2).fill=fill(bg);ws.cell(r,2).border=brd();ws.cell(r,2).alignment=aln()

        ws.cell(r,3,side_lbl).font=font(bold=True,color=tc,size=10)
        ws.cell(r,3).fill=fill(bg);ws.cell(r,3).border=brd();ws.cell(r,3).alignment=aln("center")

        ws.cell(r,4,s["ltp"]).number_format="#,##0.00"
        ws.cell(r,4).font=font(color="333333");ws.cell(r,4).fill=fill(bg)
        ws.cell(r,4).border=brd();ws.cell(r,4).alignment=aln("center")

        ws.cell(r,5,s["pChng"] and s["pChng"]/100).number_format="+0.00%;-0.00%"
        ws.cell(r,5).font=font(bold=True,color=tc);ws.cell(r,5).fill=fill(bg)
        ws.cell(r,5).border=brd();ws.cell(r,5).alignment=aln("center")

        ws.cell(r,6,s["contrib"]).number_format="+#,##0.00;-#,##0.00"
        ws.cell(r,6).font=font(bold=True,color=tc,size=13);ws.cell(r,6).fill=fill(bg)
        ws.cell(r,6).border=brd();ws.cell(r,6).alignment=aln("center")
        ws.row_dimensions[r].height = 22

    # Total contrib row
    r += 1
    total_contrib = round(sum(s["contrib"] for s in breadth["ranked_stocks"]),2)
    ws.merge_cells(f"A{r}:E{r}")
    sc(ws.cell(r,1,"TOTAL CONTRIBUTION FROM RANKED STOCKS"),
       bg="E8EAF6", fg="0D47A1", bold=True, ha="right")
    ws.cell(r,6,total_contrib).number_format="+#,##0.00;-#,##0.00"
    ws.cell(r,6).font=font(bold=True,color="0D47A1",size=13)
    ws.cell(r,6).fill=fill("E8EAF6");ws.cell(r,6).border=brd();ws.cell(r,6).alignment=aln("center")
    ws.row_dimensions[r].height = 20

    r += 2
    # ── SPECIAL WATCHLIST TABLE in Excel ─────────────────────
    ws.merge_cells(f"A{r}:G{r}")
    sc(ws.cell(r,1,"⭐  SPECIAL WATCHLIST"), bg="283593", fg="FFFFFF", bold=True, size=11, ha="center")
    r += 1
    for h, col in zip(["SYMBOL","LTP (₹)","CHANGE (₹)","% CHANGE","WEIGHT (pts)","CONTRIB (pts)",""], range(1,8)):
        sc(ws.cell(r,col,h), bg="3949AB", fg="FFFFFF", bold=True, ha="center")
    for s in special_stocks:
        if s.get("ltp") is None:
            continue
        r += 1
        pos = (s.get("pChng") or 0) >= 0
        bg  = C["pos_bg"] if pos else C["neg_bg"]
        tc  = C["pos_txt"] if pos else C["neg_txt"]
        ws.cell(r,1,s["symbol"]).font=font(bold=True,color=tc)
        ws.cell(r,1).fill=fill(bg);ws.cell(r,1).border=brd();ws.cell(r,1).alignment=aln()
        for col,val,fmt in [
            (2,s["ltp"],"#,##0.00"),
            (3,s["chng"],"+#,##0.00;-#,##0.00"),
            (4,s["pChng"] and s["pChng"]/100,"+0.00%;-0.00%"),
            (5,s.get("weight_pts"),"#,##0.00"),
            (6,s.get("contrib"),"+#,##0.00;-#,##0.00"),
        ]:
            cell=ws.cell(r,col);cell.value=val
            if val is not None and fmt: cell.number_format=fmt
            cell.font=font(bold=(col in(3,4,6)),color=tc if col in(3,4,6) else "000000")
            cell.fill=fill(bg);cell.border=brd();cell.alignment=aln("center")
        ws.row_dimensions[r].height=18

    r += 2
    ws.merge_cells(f"A{r}:D{r}")
    sc(ws.cell(r,1,"📊  INDEX SUMMARY"), bg=C["hdr_dark"], fg=C["white"], bold=True, size=11, ha="center")
    r += 1
    for h, col in zip(["INDEX","LTP (₹)","CHANGE (₹)","% CHANGE"], range(1,5)):
        sc(ws.cell(r,col,h), bg=C["hdr_mid"], fg=C["white"], bold=True, ha="center")

    for i, name in enumerate(["NIFTY 50","SENSEX","BANK NIFTY","NIFTY IT","NIFTY SMALLCAP 50"]):
        r += 1; d = indices.get(name,{}); pct = d.get("pct")
        bg = C["idx_bg"] if i%2==0 else C["idx_alt"]
        pos = pct is not None and pct >= 0
        ws.cell(r,1,name).font = font(bold=True)
        ws.cell(r,1).fill=fill(bg); ws.cell(r,1).border=brd(); ws.cell(r,1).alignment=aln()
        for col, val, fmt in [(2,d.get("ltp"),"#,##0.00"),(3,d.get("chng"),"+#,##0.00;-#,##0.00"),
                               (4,pct and pct/100,"+0.00%;-0.00%")]:
            cell = ws.cell(r,col)
            if pct is not None:
                cell.value=val; cell.number_format=fmt
                cell.font=font(bold=True, color=C["pos_txt"] if pos else C["neg_txt"])
                cell.fill=fill(C["pos_bg"] if pos else C["neg_bg"])
            else:
                cell.value="N/A"; cell.font=font(color="9E9E9E"); cell.fill=fill(bg)
            cell.border=brd(); cell.alignment=aln("center")

    sr = r + 3
    ws.merge_cells(f"A{sr}:G{sr}")
    sc(ws.cell(sr,1,f"📋  ALL NIFTY 50 STOCKS  [{len(stocks)} stocks]"),
       bg=C["hdr_dark"],fg=C["white"],bold=True,size=11,ha="center")
    sr += 1
    for h, col in zip(["#","SYMBOL","LTP (₹)","CHANGE (₹)","% CHANGE","WEIGHT (pts)","CONTRIB (pts)"],range(1,8)):
        sc(ws.cell(sr,col,h),bg="3949AB",fg=C["white"],bold=True,ha="center")

    for i, s in enumerate(stocks):
        sr += 1
        pos = (s["pChng"] or 0) >= 0
        bg  = (C["pos_bg"] if pos else C["neg_bg"]) if i%2==0 else \
              (C["pos_alt"] if pos else C["neg_alt"])
        tc  = C["pos_txt"] if pos else C["neg_txt"]
        ws.cell(sr,1,i+1).fill=fill(bg);ws.cell(sr,1).border=brd();ws.cell(sr,1).alignment=aln("center")
        ws.cell(sr,2,s["symbol"]).font=font(bold=True)
        ws.cell(sr,2).fill=fill(bg);ws.cell(sr,2).border=brd();ws.cell(sr,2).alignment=aln()
        for col,val,fmt in [(3,s["ltp"],"#,##0.00"),(4,s["chng"],"+#,##0.00;-#,##0.00"),
                             (5,s["pChng"] and s["pChng"]/100,"+0.00%;-0.00%"),
                             (6,s.get("weight_pts"),"#,##0.00"),(7,s.get("contrib"),"+0.00;-0.00")]:
            cell=ws.cell(sr,col);cell.value=val
            if val is not None and fmt: cell.number_format=fmt
            cell.font=font(bold=(col in(4,5,7)),color=tc if col in(4,5,7) else "000000")
            cell.fill=fill(bg);cell.border=brd();cell.alignment=aln("center")

    valid=[s for s in stocks if s["pChng"] is not None]
    contrib_valid=[s for s in stocks if s.get("contrib") is not None]
    top7p=sorted(contrib_valid,key=lambda x:x["contrib"],reverse=True)[:7]
    top7n=sorted(contrib_valid,key=lambda x:x["contrib"])[:7]
    trow=r+4
    for top7,col,title,hbg,tc,b1,b2 in [
        (top7p,9,"🟢  TOP 7 POSITIVE",C["grn_hdr"],C["pos_txt"],C["pos_bg"],C["pos_alt"]),
        (top7n,15,"🔴  TOP 7 NEGATIVE",C["red_hdr"],C["neg_txt"],C["neg_bg"],C["neg_alt"]),
    ]:
        ws.merge_cells(start_row=trow,start_column=col,end_row=trow,end_column=col+4)
        sc(ws.cell(trow,col,title),bg=hbg,fg=C["white"],bold=True,size=11,ha="center")
        for h,dc in zip(["SYMBOL","LTP (₹)","CHG (₹)","% CHG","CONTRIB(pts)"],range(col,col+5)):
            sc(ws.cell(trow+1,dc,h),bg=hbg,fg=C["white"],bold=True,ha="center")
        for i,s in enumerate(top7):
            tr=trow+2+i;bg=b1 if i%2==0 else b2
            for dc,val,fmt in [(col,s["symbol"],None),(col+1,s["ltp"],"#,##0.00"),
                                (col+2,s["chng"],"+#,##0.00;-#,##0.00"),
                                (col+3,s["pChng"] and s["pChng"]/100,"+0.00%;-0.00%"),
                                (col+4,s.get("contrib"),"+0.00;-0.00")]:
                cell=ws.cell(tr,dc,val)
                if val is not None and fmt: cell.number_format=fmt
                cell.font=font(bold=True,color=tc);cell.fill=fill(bg)
                cell.border=brd();cell.alignment=aln("center" if dc>col else "left")

    for col,w in {1:4,2:14,3:12,4:12,5:11,6:13,7:13,9:14,10:12,11:11,12:10,13:13,
                  15:14,16:12,17:11,18:10,19:13}.items():
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.freeze_panes="A3"
    return wb

# ════════════════════════════════════════════════════════════
# BUILD LEVELS EXCEL (File 2)
# ════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════
# EMAIL
# ════════════════════════════════════════════════════════════
def send_email(creds, snapshot_path, label, ist_dt, indices, stocks, breadth, special_stocks):
    disp    = label[:2]+":"+label[2:] if len(label)==4 else label
    subject = f"📊 NSE Snapshot {disp} IST — {ist_dt.strftime('%d %b %Y')}"

    valid   = [s for s in stocks if s.get("contrib") is not None]
    top3p   = breadth["pos_stocks"][:3]
    top3n   = breadth["neg_stocks"][:3]

    # Index HTML
    idx_html = ""
    for name in ["NIFTY 50","SENSEX","BANK NIFTY","NIFTY IT","NIFTY SMALLCAP 50"]:
        d=indices.get(name,{}); pct=d.get("pct"); ltp=d.get("ltp")
        if pct is not None:
            color="#1b5e20" if pct>=0 else "#b71c1c"
            bg="#e8f5e9" if pct>=0 else "#ffebee"
            arrow="▲" if pct>=0 else "▼"
            idx_html += (f'<tr style="background:{bg}"><td style="padding:7px 14px;font-weight:bold">{name}</td>'
                f'<td style="padding:7px;color:{color};font-weight:bold;text-align:center">{arrow} {pct:+.2f}%</td>'
                f'<td style="padding:7px;text-align:center">₹{ltp:,.2f}</td></tr>')
        else:
            idx_html += f'<tr><td style="padding:7px 14px;font-weight:bold">{name}</td><td style="padding:7px;color:#9e9e9e;text-align:center">N/A</td><td>—</td></tr>'

    def srows(lst, color, bg):
        return "".join(
            f'<tr style="background:{bg}"><td style="padding:6px 12px;font-weight:bold">{s["symbol"]}</td>'
            f'<td style="padding:6px;color:{color};font-weight:bold;text-align:center">{s["pChng"]:+.2f}%</td>'
            f'<td style="padding:6px;color:{color};text-align:center">{s["contrib"]:+.2f} pts</td></tr>'
            for s in lst)

    note = ('<div style="background:#fff3e0;border-radius:6px;padding:10px 14px;margin-top:12px;'
            'font-size:13px;color:#e65100">⚠️ Market closed or data unavailable.</div>'
            if not valid else "")

    html = f"""<html><body style="font-family:Arial,sans-serif;max-width:640px;margin:auto">
<div style="background:#0d47a1;color:white;padding:18px 22px;border-radius:10px 10px 0 0">
  <h2 style="margin:0;font-size:20px">📊 NSE Snapshot — {disp} IST</h2>
  <p style="margin:5px 0 0;opacity:.85;font-size:13px">{ist_dt.strftime('%d %b %Y  %H:%M:%S IST')}</p>
</div>
<div style="border:1px solid #ddd;border-top:none;padding:18px;border-radius:0 0 10px 10px">

  <!-- BREADTH SCORE BOX -->
  <div style="background:#e8eaf6;border:1px solid #3949ab;border-radius:10px;padding:16px;margin-bottom:16px;text-align:center">
    <div style="font-size:13px;color:#3949ab;font-weight:500;margin-bottom:12px">📊 NIFTY 50 BREADTH SCORE</div>

    <!-- PRIMARY DISPLAY: Positive Score | Negative Score -->
    <div style="background:#0d47a1;border-radius:10px;padding:16px;margin-bottom:14px;display:inline-block;min-width:280px">
      <div style="font-size:13px;color:#90caf9;margin-bottom:6px">Highest Positive | Lowest Negative</div>
      <div style="display:flex;justify-content:center;align-items:center;gap:8px">
        <div style="font-size:52px;font-weight:900;color:#69f0ae;line-height:1">{breadth["pos_score"]}</div>
        <div style="font-size:36px;font-weight:bold;color:white;margin:0 4px">|</div>
        <div style="font-size:52px;font-weight:900;color:#ff5252;line-height:1">{breadth["neg_score"]}</div>
      </div>
      <div style="font-size:12px;color:#90caf9;margin-top:6px">Total = {breadth["total_score"]} (always 7)</div>
    </div>

    <!-- DETAIL ROW -->
    <div style="display:flex;justify-content:center;align-items:stretch;gap:10px;margin-top:4px">
      <div style="background:#e8f5e9;border:1px solid #2e7d32;border-radius:8px;padding:10px 18px;flex:1">
        <div style="font-size:11px;color:#2e7d32;font-weight:500">🟢 POSITIVE STOCKS</div>
        <div style="font-size:22px;font-weight:bold;color:#1b5e20">{breadth["pos_count"]} stocks</div>
        <div style="font-size:12px;color:#1b5e20;font-weight:500">{breadth["pos_contrib_sum"]:+.2f} pts total</div>
        <div style="font-size:11px;color:#555">Score = {breadth["pos_count"]} × 7 ÷ 50 = {breadth["pos_score"]}</div>
      </div>
      <div style="background:#ffebee;border:1px solid #c62828;border-radius:8px;padding:10px 18px;flex:1">
        <div style="font-size:11px;color:#c62828;font-weight:500">🔴 NEGATIVE STOCKS</div>
        <div style="font-size:22px;font-weight:bold;color:#b71c1c">{breadth["neg_count"]} stocks</div>
        <div style="font-size:12px;color:#b71c1c;font-weight:500">{breadth["neg_contrib_sum"]:+.2f} pts total</div>
        <div style="font-size:11px;color:#555">Score = {breadth["neg_count"]} × 7 ÷ 50 = {breadth["neg_score"]}</div>
      </div>
    </div>
  </div>

  <!-- RANKED BREADTH TABLE — higher score always first -->
  <div style="margin-bottom:16px">
    <div style="font-size:13px;color:#37474f;font-weight:600;margin-bottom:8px;text-align:center">
      📋 BREADTH RANKING TABLE &nbsp;·&nbsp; Higher score listed first
    </div>
    <table width="100%" cellspacing="0" style="border-collapse:collapse;border-radius:10px;overflow:hidden;border:2px solid #0d47a1">
      <tr style="background:#0d47a1;color:white">
        <th style="padding:10px 14px;text-align:center;font-size:13px">RANK</th>
        <th style="padding:10px 14px;text-align:center;font-size:13px">SIDE</th>
        <th style="padding:10px 14px;text-align:center;font-size:13px">SCORE</th>
        <th style="padding:10px 14px;text-align:center;font-size:13px">STOCKS</th>
        <th style="padding:10px 14px;text-align:center;font-size:13px">CONTRIB SUM (pts)</th>
        <th style="padding:10px 14px;text-align:center;font-size:13px">FORMULA</th>
      </tr>
      <tr style="background:#{breadth["first_bg"]}">
        <td style="padding:12px 14px;text-align:center;font-size:18px;font-weight:900;color:#{breadth["first_color"]}">1st</td>
        <td style="padding:12px 14px;text-align:center;font-size:14px;font-weight:bold;color:#{breadth["first_color"]}">{breadth["first_label"]}</td>
        <td style="padding:12px 14px;text-align:center;font-size:32px;font-weight:900;color:#{breadth["first_color"]}">{breadth["first_score"]}</td>
        <td style="padding:12px 14px;text-align:center;font-size:16px;font-weight:bold;color:#{breadth["first_color"]}">{breadth["first_count"]}</td>
        <td style="padding:12px 14px;text-align:center;font-size:14px;font-weight:bold;color:#{breadth["first_color"]}">{breadth["first_pts"]:+.2f}</td>
        <td style="padding:12px 14px;text-align:center;font-size:12px;color:#555">{breadth["first_count"]}×7÷50={breadth["first_score"]}</td>
      </tr>
      <tr style="background:#{breadth["second_bg"]}">
        <td style="padding:12px 14px;text-align:center;font-size:18px;font-weight:900;color:#{breadth["second_color"]}">2nd</td>
        <td style="padding:12px 14px;text-align:center;font-size:14px;font-weight:bold;color:#{breadth["second_color"]}">{breadth["second_label"]}</td>
        <td style="padding:12px 14px;text-align:center;font-size:32px;font-weight:900;color:#{breadth["second_color"]}">{breadth["second_score"]}</td>
        <td style="padding:12px 14px;text-align:center;font-size:16px;font-weight:bold;color:#{breadth["second_color"]}">{breadth["second_count"]}</td>
        <td style="padding:12px 14px;text-align:center;font-size:14px;font-weight:bold;color:#{breadth["second_color"]}">{breadth["second_pts"]:+.2f}</td>
        <td style="padding:12px 14px;text-align:center;font-size:12px;color:#555">{breadth["second_count"]}×7÷50={breadth["second_score"]}</td>
      </tr>
      <tr style="background:#e8eaf6">
        <td colspan="2" style="padding:10px 14px;text-align:center;font-size:13px;font-weight:bold;color:#0d47a1">TOTAL</td>
        <td style="padding:10px 14px;text-align:center;font-size:24px;font-weight:900;color:#0d47a1">{breadth["total_score"]}</td>
        <td style="padding:10px 14px;text-align:center;font-size:13px;color:#555">{breadth["pos_count"]+breadth["neg_count"]} stocks</td>
        <td style="padding:10px 14px;text-align:center;font-size:13px;font-weight:bold;color:#0d47a1">{round(breadth["pos_contrib_sum"]+breadth["neg_contrib_sum"],2):+.2f}</td>
        <td style="padding:10px 14px;text-align:center;font-size:12px;color:#0d47a1">Always = 7</td>
      </tr>
    </table>
    <div style="text-align:center;margin-top:6px;font-size:12px;color:#666">
      🏆 Dominant side: <b style="color:#{breadth["first_color"]}">{breadth["dominant"]}</b>
      &nbsp;·&nbsp; Ranked display: <b>{breadth["ranked_display"]}</b>
    </div>
  </div>

  <!-- 7-STOCK RANKED TABLE -->
  <div style="margin-bottom:16px">
    <div style="font-size:13px;color:#37474f;font-weight:600;margin-bottom:8px;text-align:center">
      🏆 TOP {breadth["total_score"]} STOCKS BY BREADTH SCORE
      &nbsp;·&nbsp;
      {breadth["first_score"]} {breadth["first_label"]} + {breadth["second_score"]} {breadth["second_label"]}
    </div>
    <table width="100%" cellspacing="0" style="border-collapse:collapse;border:2px solid #0d47a1;border-radius:10px;overflow:hidden">
      <tr style="background:#0d47a1;color:white">
        <th style="padding:8px 10px;text-align:center">#</th>
        <th style="padding:8px 10px;text-align:left">SYMBOL</th>
        <th style="padding:8px 10px;text-align:center">SIDE</th>
        <th style="padding:8px 10px;text-align:right">LTP (₹)</th>
        <th style="padding:8px 10px;text-align:right">% CHANGE</th>
        <th style="padding:8px 10px;text-align:right">CONTRIBUTION (pts)</th>
      </tr>
      {"".join(
        f'<tr style="background:#{s["bg"]}">' +
        f'<td style="padding:8px 10px;text-align:center;font-weight:bold;color:#{s["color"]}">{s["rank"]}</td>' +
        f'<td style="padding:8px 10px;font-weight:bold;color:#{s["color"]};font-size:14px">{s["symbol"]}</td>' +
        f'<td style="padding:8px 10px;text-align:center;font-size:11px;font-weight:600;color:#{s["color"]}">{"🟢 POS" if s["side"]=="POSITIVE" else "🔴 NEG"}</td>' +
        f'<td style="padding:8px 10px;text-align:right;color:#333">{s["ltp"]:,.2f}</td>' +
        f'<td style="padding:8px 10px;text-align:right;font-weight:bold;color:#{s["color"]}">{s["pChng"]:+.2f}%</td>' +
        f'<td style="padding:8px 10px;text-align:right;font-weight:bold;color:#{s["color"]};font-size:14px">{s["contrib"]:+.2f}</td>' +
        '</tr>'
        for s in breadth["ranked_stocks"]
      )}
      <tr style="background:#e8eaf6">
        <td colspan="5" style="padding:8px 10px;text-align:right;font-weight:bold;color:#0d47a1">TOTAL CONTRIBUTION</td>
        <td style="padding:8px 10px;text-align:right;font-weight:bold;color:#0d47a1;font-size:14px">
          {round(sum(s["contrib"] for s in breadth["ranked_stocks"]),2):+.2f} pts
        </td>
      </tr>
    </table>
  </div>

  <!-- SPECIAL WATCHLIST TABLE -->
  <h3 style="color:#0d47a1;margin:0 0 8px">⭐ Special Watchlist</h3>
  <table width="100%" cellspacing="0" style="border-collapse:collapse;border:1px solid #e0e0e0;margin-bottom:16px">
    <tr style="background:#283593;color:white">
      <th style="padding:7px 10px;text-align:left">SYMBOL</th>
      <th style="padding:7px 10px;text-align:right">LTP (₹)</th>
      <th style="padding:7px 10px;text-align:right">CHANGE (₹)</th>
      <th style="padding:7px 10px;text-align:right">% CHANGE</th>
      <th style="padding:7px 10px;text-align:right">CONTRIB (pts)</th>
    </tr>
    {"".join(
      f'<tr style="background:{"#e8f5e9" if (s.get("pChng") or 0)>=0 else "#ffebee"}">' +
      f'<td style="padding:7px 10px;font-weight:bold;color:{"#1b5e20" if (s.get("pChng") or 0)>=0 else "#b71c1c"}">{s["symbol"]}</td>' +
      f'<td style="padding:7px 10px;text-align:right">₹{s["ltp"]:,.2f}</td>' +
      f'<td style="padding:7px 10px;text-align:right;color:{"#1b5e20" if (s.get("chng") or 0)>=0 else "#b71c1c"};font-weight:bold">{s["chng"]:+.2f}</td>' +
      f'<td style="padding:7px 10px;text-align:right;color:{"#1b5e20" if (s.get("pChng") or 0)>=0 else "#b71c1c"};font-weight:bold">{s["pChng"]:+.2f}%</td>' +
      f'<td style="padding:7px 10px;text-align:right;color:{"#1b5e20" if (s.get("contrib") or 0)>=0 else "#b71c1c"};font-weight:bold">{s["contrib"]:+.2f} pts</td>' +
      '</tr>'
      for s in special_stocks if s.get("ltp") is not None
    )}
  </table>

  <h3 style="color:#0d47a1;margin:0 0 8px">Index Performance</h3>
  <table width="100%" cellspacing="0" style="border-collapse:collapse;border:1px solid #e0e0e0">
    <tr style="background:#1976d2;color:white">
      <th style="padding:8px 14px;text-align:left">Index</th>
      <th style="padding:8px">% Change</th><th style="padding:8px">LTP</th>
    </tr>{idx_html}
  </table>

  <table width="100%" style="margin-top:16px;border-collapse:collapse"><tr valign="top">
    <td width="50%" style="padding-right:8px">
      <h3 style="color:#2e7d32;margin:0 0 8px">🏆 Top 3 Boosters</h3>
      <table width="100%" cellspacing="0" style="border-collapse:collapse;border:1px solid #e0e0e0">
        {srows(top3p,"#1b5e20","#e8f5e9") if top3p else '<tr><td style="padding:8px;color:#9e9e9e">No data</td></tr>'}
      </table>
    </td>
    <td width="50%" style="padding-left:8px">
      <h3 style="color:#c62828;margin:0 0 8px">📉 Top 3 Draggers</h3>
      <table width="100%" cellspacing="0" style="border-collapse:collapse;border:1px solid #e0e0e0">
        {srows(top3n,"#b71c1c","#ffebee") if top3n else '<tr><td style="padding:8px;color:#9e9e9e">No data</td></tr>'}
      </table>
    </td>
  </tr></table>
  {note}
  <div style="margin-top:16px;padding:12px 14px;background:#f5f5f5;border-radius:8px;font-size:13px">
    📎 Excel with all 50 stocks attached<br>
    ☁️ Saved to Google Drive → NSE Snapshots folder
  </div>
</div></body></html>"""

    msg = MIMEMultipart("alternative")
    msg["From"]=creds["gmail_sender"]; msg["To"]=creds["recipient"]; msg["Subject"]=subject
    msg.attach(MIMEText(html,"html"))
    with open(snapshot_path,"rb") as f:
        part=MIMEBase("application","vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read()); encoders.encode_base64(part)
        part.add_header("Content-Disposition","attachment",filename=os.path.basename(snapshot_path))
        msg.attach(part)
    with smtplib.SMTP_SSL("smtp.gmail.com",465) as s:
        s.login(creds["gmail_sender"],creds["gmail_pass"])
        s.sendmail(creds["gmail_sender"],creds["recipient"],msg.as_string())
    print(f"✅ Email sent to {creds['recipient']}")


def upload_drive(creds, xlsx_path):
    folder_id = creds["drive_folder"]
    if not folder_id or len(folder_id) < 10:
        print("❌ Drive folder ID missing"); return
    try:
        svc = build("drive","v3",credentials=get_sa_creds(
            creds["sa_json"],["https://www.googleapis.com/auth/drive"]))
        name = os.path.basename(xlsx_path)
        q    = f"name='{name}' and '{folder_id}' in parents and trashed=false"
        for f in svc.files().list(q=q,fields="files(id)",supportsAllDrives=True,
                includeItemsFromAllDrives=True).execute().get("files",[]):
            svc.files().delete(fileId=f["id"],supportsAllDrives=True).execute()
        meta  = {"name":name,"parents":[folder_id]}
        media = MediaFileUpload(xlsx_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=False)
        res = svc.files().create(body=meta,media_body=media,
            fields="id,webViewLink",supportsAllDrives=True).execute()
        print(f"✅ Drive: {name}  →  {res.get('webViewLink','')}")
    except Exception as e:
        print(f"❌ Drive upload failed: {e}")

# ════════════════════════════════════════════════════════════
# GOOGLE SHEETS
# ════════════════════════════════════════════════════════════
def update_sheets(creds, label, ist_dt, indices, stocks, breadth, special_stocks):
    try:
        gc  = gspread.authorize(get_sa_creds(
            creds["sa_json"],["https://www.googleapis.com/auth/spreadsheets"]))
        sh  = gc.open_by_key(creds["sheet_id"])
        disp = label[:2]+":"+label[2:] if len(label)==4 else label

        tab1 = "Snapshot_" + label
        try: sh.del_worksheet(sh.worksheet(tab1))
        except: pass
        ws1 = sh.add_worksheet(title=tab1, rows=200, cols=20)

        rows1 = []

        # ── Row 1-2: Title ────────────────────────────────────
        rows1.append([f"NSE Market Snapshot — {disp} IST"] + [""]*19)
        rows1.append([f"Captured: {ist_dt.strftime('%d-%b-%Y %H:%M:%S IST')}"] + [""]*19)
        rows1.append([""]*20)

        # ── Row 4: Breadth Score Summary ─────────────────────
        rows1.append([
            "📊 BREADTH SCORE",
            f"🟢 POSITIVE: {breadth['pos_score']}",
            f"({breadth['pos_count']} stocks)",
            f"{breadth['pos_contrib_sum']:+.2f} pts",
            "|",
            f"🔴 NEGATIVE: {breadth['neg_score']}",
            f"({breadth['neg_count']} stocks)",
            f"{breadth['neg_contrib_sum']:+.2f} pts",
            "=",
            f"TOTAL: {breadth['total_score']}",
            "always 7",
            "", "", "", "", "", "", "", "", ""
        ])
        rows1.append([""]*20)

        # ── Row 6: Ranked Breadth Table Header ───────────────
        rows1.append([
            "RANK", "SIDE", "SCORE", "STOCKS COUNT",
            "CONTRIB SUM (pts)", "FORMULA", "", "", "", "",
            "", "", "", "", "", "", "", "", "", ""
        ])

        # First ranked row (dominant)
        rows1.append([
            "1st",
            breadth["first_label"],
            breadth["first_score"],
            breadth["first_count"],
            f"{breadth['first_pts']:+.2f}",
            f"{breadth['first_count']}x7/50={breadth['first_score']}",
            "", "", "", "", "", "", "", "", "", "", "", "", "", ""
        ])

        # Second ranked row
        rows1.append([
            "2nd",
            breadth["second_label"],
            breadth["second_score"],
            breadth["second_count"],
            f"{breadth['second_pts']:+.2f}",
            f"{breadth['second_count']}x7/50={breadth['second_score']}",
            "", "", "", "", "", "", "", "", "", "", "", "", "", ""
        ])

        # Total row
        total_contrib = round(breadth["pos_contrib_sum"] + breadth["neg_contrib_sum"], 2)
        rows1.append([
            "TOTAL",
            f"Dominant: {breadth['dominant']}",
            breadth["total_score"],
            breadth["pos_count"] + breadth["neg_count"],
            f"{total_contrib:+.2f}",
            f"Ranked: {breadth['ranked_display']}",
            "", "", "", "", "", "", "", "", "", "", "", "", "", ""
        ])
        rows1.append([""]*20)

        # ── 7-Stock Ranked Table Header ───────────────────────
        rows1.append([
            f"🏆 TOP {breadth['total_score']} STOCKS BY BREADTH SCORE",
            f"{breadth['first_score']} {breadth['first_label']} + {breadth['second_score']} {breadth['second_label']}",
            "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""
        ])
        rows1.append([
            "#", "SYMBOL", "SIDE", "LTP (₹)",
            "% CHANGE", "CONTRIBUTION (pts)", "",
            "", "", "", "", "", "", "", "", "", "", "", "", ""
        ])

        # 7 ranked stocks
        for s in breadth["ranked_stocks"]:
            side_lbl = "🟢 POSITIVE" if s["side"] == "POSITIVE" else "🔴 NEGATIVE"
            rows1.append([
                s["rank"],
                s["symbol"],
                side_lbl,
                s["ltp"],
                f"{s['pChng']:+.2f}%",
                f"{s['contrib']:+.2f}",
                "", "", "", "", "", "", "", "", "", "", "", "", "", ""
            ])

        # Total contrib of ranked stocks
        ranked_total = round(sum(s["contrib"] for s in breadth["ranked_stocks"]), 2)
        rows1.append([
            "TOTAL", "", "", "",
            "Total Contribution:",
            f"{ranked_total:+.2f} pts",
            "", "", "", "", "", "", "", "", "", "", "", "", "", ""
        ])
        rows1.append([""]*20)

        # ── Special Watchlist ─────────────────────────────────
        rows1.append(["⭐ SPECIAL WATCHLIST"] + [""]*19)
        rows1.append(["SYMBOL","LTP","CHANGE","% CHANGE","WEIGHT(pts)","CONTRIB(pts)"] + [""]*14)
        for s in special_stocks:
            if s.get("ltp") is None:
                continue
            rows1.append([
                s["symbol"],
                s["ltp"],
                f'{s["chng"]:+.2f}' if s["chng"] is not None else "",
                f'{s["pChng"]:+.2f}%' if s["pChng"] is not None else "",
                f'{s["weight_pts"]:.2f}' if s.get("weight_pts") is not None else "",
                f'{s["contrib"]:+.2f}' if s.get("contrib") is not None else "",
            ] + [""]*14)
        rows1.append([""]*20)

        # ── Index Summary ─────────────────────────────────────
        rows1.append(["📊 INDEX SUMMARY", "LTP", "CHANGE", "% CHANGE"] + [""]*16)
        for name in ["NIFTY 50","SENSEX","BANK NIFTY","NIFTY IT","NIFTY SMALLCAP 50"]:
            d=indices.get(name,{}); p=d.get("pct")
            rows1.append([
                name,
                d.get("ltp","N/A"),
                d.get("chng","N/A"),
                f"{p:+.2f}%" if p is not None else "N/A"
            ] + [""]*16)
        rows1.append([""]*20)

        # ── All 50 Stocks ─────────────────────────────────────
        rows1.append([
            "#","SYMBOL","LTP","CHG","% CHG","WT(pts)","CONTRIB","",
            "TOP POSITIVE","LTP","CHG","% CHG","CONTRIB","",
            "TOP NEGATIVE","LTP","CHG","% CHG","CONTRIB",""
        ])

        top7p = breadth["pos_stocks"][:7]
        top7n = breadth["neg_stocks"][:7]

        for i, s in enumerate(stocks):
            p = s["pChng"]
            row = [
                i+1, s["symbol"], s.get("ltp",""),
                f'{s["chng"]:+.2f}' if s["chng"] is not None else "",
                f'{p:+.2f}%' if p is not None else "",
                f'{s["weight_pts"]:.2f}' if s.get("weight_pts") is not None else "",
                f'{s["contrib"]:+.2f}' if s.get("contrib") is not None else "",
                ""
            ]
            row += ([
                top7p[i]["symbol"], top7p[i]["ltp"],
                f'{top7p[i]["chng"]:+.2f}', f'{top7p[i]["pChng"]:+.2f}%',
                f'{top7p[i]["contrib"]:+.2f}', ""
            ] if i < len(top7p) else [""]*6)
            row += ([
                top7n[i]["symbol"], top7n[i]["ltp"],
                f'{top7n[i]["chng"]:+.2f}', f'{top7n[i]["pChng"]:+.2f}%',
                f'{top7n[i]["contrib"]:+.2f}'
            ] if i < len(top7n) else [""]*5)
            rows1.append(row)

        ws1.update("A1", rows1)
        print(f"✅ Google Sheets updated: {tab1}")

    except Exception as e:
        print(f"⚠️  Sheets failed: {e}")
        traceback.print_exc()


def main():
    ist_dt    = datetime.now(IST)
    label     = env("MANUAL_LABEL") or ist_dt.strftime("%H%M")
    disp      = label[:2]+":"+label[2:] if len(label)==4 else label

    print(f"\n{'='*60}")
    print(f"  NSE Snapshot — {disp} IST | {ist_dt.strftime('%d-%b-%Y')}")
    print(f"{'='*60}")

    creds      = get_creds()
    obj        = angel_login(creds)
    token_map  = load_token_map()
    indices    = fetch_indices(obj)

    nifty_ltp  = indices.get("NIFTY 50",{}).get("ltp")   or 0
    prev_close = indices.get("NIFTY 50",{}).get("close") or 0
    print(f"\n  Nifty LTP: {nifty_ltp}  |  Prev Close: {prev_close}")

    stocks = fetch_nifty50(obj, nifty_ltp, token_map)

    print("\n📊 Calculating breadth score...")
    breadth = calc_breadth(stocks)
    special_stocks = get_special_stocks(stocks)
    print(f"  ✅ Special watchlist: {len(special_stocks)} stocks loaded")
    print(f"  Positive: {breadth['pos_count']} stocks → score {breadth['pos_score']}")
    print(f"  Negative: {breadth['neg_count']} stocks → score {breadth['neg_score']}")
    print(f"  Dominant: {breadth['dominant']}  |  Ranked: {breadth['ranked_display']}")
    print(f"  Total score: {breadth['total_score']} (always 7)")

    os.makedirs("output", exist_ok=True)
    date_str = ist_dt.strftime("%Y-%m-%d")

    print("\n📁 Building Excel...")
    wb1 = build_snapshot_excel(label, ist_dt, indices, stocks, breadth, special_stocks)
    snap_path = f"output/NSE_{date_str}_{label}.xlsx"
    wb1.save(snap_path)
    print(f"  ✅ {snap_path}")

    print("\n📧 Sending email...")
    send_email(creds, snap_path, label, ist_dt, indices, stocks, breadth, special_stocks)

    print("\n☁️  Uploading to Drive...")
    upload_drive(creds, snap_path)

    print("\n📊 Updating Sheets...")
    update_sheets(creds, label, ist_dt, indices, stocks, breadth, special_stocks)

    print(f"\n✅ ALL DONE — {disp} IST\n")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}")
        traceback.print_exc()
        raise
