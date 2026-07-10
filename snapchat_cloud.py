"""
snapshot_cloud.py v10
Angel One Smart API → Excel (Nifty 50 snapshot) + Breadth Score + Email + Drive + Sheets
"""

import os, json, pyotp, gspread, traceback, smtplib, time, requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from SmartApi import SmartConnect
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

IST = ZoneInfo("Asia/Kolkata")
INSTRUMENT_MASTER_URL = "https://margincalculator.angelone.in/OpenAPI_File/files/OpenAPIScripMaster.json"

# ════════════════════════════════════════════════════════════
# BREADTH SCORE FORMULA: count * 7 / 50  → round to int
# Positive score + Negative score = 7 always
# ════════════════════════════════════════════════════════════
def breadth_score(count):
    return round(count * 7 / 50)

def calc_breadth(stocks):
    """Calculate breadth score from stock contribution data."""
    valid   = [s for s in stocks if s.get("contrib") is not None]
    pos     = [s for s in valid if s["contrib"] >= 0]
    neg     = [s for s in valid if s["contrib"] <  0]
    pos_cnt = len(pos)
    neg_cnt = len(neg)
    pos_pts = round(sum(s["contrib"] for s in pos), 2)
    neg_pts = round(sum(s["contrib"] for s in neg), 2)
    pos_sc  = breadth_score(pos_cnt)
    neg_sc  = breadth_score(neg_cnt)
    total   = pos_sc + neg_sc
    return {
        "pos_count": pos_cnt,
        "neg_count": neg_cnt,
        "pos_contrib_sum": pos_pts,
        "neg_contrib_sum": neg_pts,
        "pos_score":  pos_sc,
        "neg_score":  neg_sc,
        "total_score": total,
        "pos_stocks": sorted(pos, key=lambda x: x["contrib"], reverse=True),
        "neg_stocks": sorted(neg, key=lambda x: x["contrib"]),
    }

# ════════════════════════════════════════════════════════════
# CREDENTIALS
# ════════════════════════════════════════════════════════════
def env(key, default=""):
    return os.environ.get(key, default).strip()

def get_creds():
    return {
        "api_key":      env("ANGEL_API_KEY"),
        "client_id":    env("ANGEL_CLIENT_ID"),
        "password":     env("ANGEL_PASSWORD"),
        "totp_secret":  env("ANGEL_TOTP_SECRET"),
        "gmail_sender": env("GMAIL_SENDER"),
        "gmail_pass":   env("GMAIL_APP_PASSWORD"),
        "recipient":    env("RECIPIENT_EMAIL"),
        "drive_folder": env("GOOGLE_DRIVE_FOLDER_ID"),
        "sheet_id":     env("GOOGLE_SHEET_ID"),
        "sa_json":      env("SERVICE_ACCOUNT_JSON"),

    }

def get_sa_creds(sa_json_str, scopes):
    return Credentials.from_service_account_info(json.loads(sa_json_str), scopes=scopes)

# ════════════════════════════════════════════════════════════
# EXCEL STYLES
# ════════════════════════════════════════════════════════════
C = {
    "title_bg":"0D47A1","hdr_mid":"1976D2","hdr_dark":"283593",
    "pos_bg":"E8F5E9","pos_alt":"F1F8E9","pos_txt":"1B5E20",
    "neg_bg":"FFEBEE","neg_alt":"FCE4EC","neg_txt":"B71C1C",
    "idx_bg":"E8EAF6","idx_alt":"FFFFFF",
    "grn_hdr":"2E7D32","red_hdr":"C62828","white":"FFFFFF",
    "yellow":"FFF176","yellow_lt":"FFF9C4",
}
def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
def font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")
def aln(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def brd():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)
def sc(cell, bg=None, fg="000000", bold=False, size=10, ha="left"):
    if bg: cell.fill = fill(bg)
    cell.font = font(bold, fg, size)
    cell.alignment = aln(ha)
    cell.border = brd()

# ════════════════════════════════════════════════════════════
# ANGEL ONE
# ════════════════════════════════════════════════════════════
def angel_login(creds):
    print("🔑 Logging in to Angel One...")
    obj  = SmartConnect(api_key=creds["api_key"])
    totp = pyotp.TOTP(creds["totp_secret"]).now()
    data = obj.generateSession(creds["client_id"], creds["password"], totp)
    if data["status"] is False:
        raise Exception("Login failed: " + str(data.get("message","")))
    print("✅ Login OK")
    return obj

def load_token_map():
    print("\n📥 Downloading instrument master...")
    r = requests.get(INSTRUMENT_MASTER_URL, timeout=30)
    r.raise_for_status()
    data = r.json()
    token_map = {}
    for item in data:
        if item.get("exch_seg") == "NSE" and item.get("symbol","").endswith("-EQ"):
            base = item["symbol"].replace("-EQ","")
            token_map[base] = {"token": item["token"], "tradingsymbol": item["symbol"]}
    print(f"  ✅ {len(token_map)} NSE equities loaded")
    return token_map

INDEX_TOKENS = {
    "NIFTY 50":       ("NSE", "Nifty 50",          "99926000"),
    "SENSEX":         ("BSE", "SENSEX",             "99919000"),
    "BANK NIFTY":     ("NSE", "Nifty Bank",         "99926009"),
    "NIFTY IT":       ("NSE", "Nifty IT",           "99926011"),
    "NIFTY SMALLCAP 50": ("NSE", "Nifty Smallcap 50", "99926061"),
}

def fetch_one_index(obj, name, exch, sym, token):
    try:
        r = obj.ltpData(exch, sym, token)
        if r.get("status") and r.get("data"):
            ltp   = float(r["data"].get("ltp",   0) or 0)
            close = float(r["data"].get("close", ltp) or ltp)
            chng  = round(ltp - close, 2)
            pct   = round((chng / close) * 100, 2) if close else 0.0
            return name, {"ltp": ltp, "chng": chng, "pct": pct, "close": close}
        return name, {"ltp": None, "chng": None, "pct": None, "close": None}
    except Exception as e:
        print(f"  ❌ {name}: {e}")
        return name, {"ltp": None, "chng": None, "pct": None, "close": None}

def fetch_indices(obj):
    print("\n📊 Fetching indices in parallel...")
    result = {}
    with ThreadPoolExecutor(max_workers=5) as ex:
        futures = {ex.submit(fetch_one_index, obj, name, exch, sym, token): name
                   for name, (exch, sym, token) in INDEX_TOKENS.items()}
        for future in as_completed(futures):
            name, data = future.result()
            result[name] = data
            pct = data.get("pct")
            ltp = data.get("ltp")
            if ltp:
                print(f"  ✅ {name}: {ltp}  {pct:+.2f}%")
            else:
                print(f"  ⚠️  {name}: no data")
    return result

NIFTY50_LIST = {
    "RELIANCE":9.25,"HDFCBANK":12.80,"ICICIBANK":8.85,"INFY":5.85,"TCS":3.90,
    "BHARTIARTL":4.20,"ITC":3.55,"LT":3.45,"KOTAKBANK":2.95,"AXISBANK":2.85,
    "SBIN":2.75,"HINDUNILVR":2.30,"BAJFINANCE":2.05,"M&M":1.95,"MARUTI":1.55,
    "SUNPHARMA":1.50,"ULTRACEMCO":1.35,"HCLTECH":1.30,"TITAN":1.25,"BAJAJFINSV":1.20,
    "ASIANPAINT":1.15,"ADANIENT":1.10,"NTPC":1.10,"ONGC":1.05,"POWERGRID":1.05,
    "TATASTEEL":1.00,"TECHM":0.95,"WIPRO":0.90,"ADANIPORTS":0.90,"COALINDIA":0.85,
    "JSWSTEEL":0.80,"BAJAJ-AUTO":0.80,"GRASIM":0.70,"DRREDDY":0.65,"CIPLA":0.65,
    "EICHERMOT":0.65,"INDUSINDBK":0.60,"APOLLOHOSP":0.55,"TATACONSUM":0.55,
    "BEL":0.50,"SHRIRAMFIN":0.45,"HINDALCO":0.75,"HDFCLIFE":0.90,"INDIGO":0.70,
    "JIOFIN":0.85,"MAXHEALTH":0.55,"SBILIFE":0.85,"TMPV":0.60,"TRENT":0.95,
    "ETERNAL":0.95,
}

def fetch_one_stock(obj, sym, wt, info, nifty_ltp):
    """Fetch a single stock — called in parallel."""
    if not info:
        wpts = round(nifty_ltp*(wt/100),2) if nifty_ltp else None
        return {"symbol":sym,"ltp":None,"chng":None,"pChng":None,
                "weight":wt,"weight_pts":wpts,"contrib":None}
    for attempt in range(3):
        try:
            r = obj.ltpData("NSE", info["tradingsymbol"], info["token"])
            if r.get("status") and r.get("data"):
                ltp   = float(r["data"].get("ltp",   0) or 0)
                close = float(r["data"].get("close", ltp) or ltp)
                chng  = round(ltp - close, 2)
                pct   = round((chng / close) * 100, 2) if close else 0.0
                contrib   = round(nifty_ltp * (pct/100) * (wt/100), 2) if nifty_ltp else None
                weight_pts= round(nifty_ltp * (wt/100), 2) if nifty_ltp else None
                return {"symbol":sym,"ltp":ltp,"chng":chng,"pChng":pct,
                        "weight":wt,"weight_pts":weight_pts,"contrib":contrib}
            else:
                msg = r.get("message","")
                if "rate" in msg.lower() and attempt < 2:
                    time.sleep(2); continue
                wpts = round(nifty_ltp*(wt/100),2) if nifty_ltp else None
                return {"symbol":sym,"ltp":None,"chng":None,"pChng":None,
                        "weight":wt,"weight_pts":wpts,"contrib":None}
        except Exception as e:
            if "rate" in str(e).lower() and attempt < 2:
                time.sleep(3); continue
            wpts = round(nifty_ltp*(wt/100),2) if nifty_ltp else None
            return {"symbol":sym,"ltp":None,"chng":None,"pChng":None,
                    "weight":wt,"weight_pts":wpts,"contrib":None}
    wpts = round(nifty_ltp*(wt/100),2) if nifty_ltp else None
    return {"symbol":sym,"ltp":None,"chng":None,"pChng":None,
            "weight":wt,"weight_pts":wpts,"contrib":None}

def fetch_nifty50(obj, nifty_ltp, token_map):
    print("\n📋 Fetching Nifty 50 stocks in parallel (10 at a time)...")
    symbols = list(NIFTY50_LIST.items())
    results = {}

    # Fetch in batches of 10 parallel — avoids rate limit
    BATCH = 10
    for i in range(0, len(symbols), BATCH):
        batch = symbols[i:i+BATCH]
        with ThreadPoolExecutor(max_workers=BATCH) as ex:
            futures = {ex.submit(fetch_one_stock, obj, sym, wt,
                                 token_map.get(sym), nifty_ltp): sym
                       for sym, wt in batch}
            for future in as_completed(futures):
                s = future.result()
                results[s["symbol"]] = s
                if s["ltp"]:
                    print(f"  ✅ {s['symbol']}: {s['ltp']}  {s['pChng']:+.2f}%")
        # Small pause between batches to avoid rate limit
        if i + BATCH < len(symbols):
            time.sleep(1)

    # Return in original NIFTY50_LIST order
    stocks = [results[sym] for sym, _ in symbols if sym in results]
    filled = sum(1 for s in stocks if s["ltp"] is not None)
    print(f"\n  Result: {filled}/{len(stocks)} stocks with data")
    return stocks

# ════════════════════════════════════════════════════════════
# LEVEL CALCULATOR
# ════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════
# BREADTH SCORE
# Formula: count * 7 / 50 → round to integer
# Positive score + Negative score always = 7
# ════════════════════════════════════════════════════════════
def breadth_score(count):
    return round(count * 7 / 50)

def calc_breadth(stocks):
    """
    Separates stocks into positive and negative by contribution pts.
    Calculates sum of contribution pts for each side.
    Applies breadth score formula: count * 7 / 50 → rounded integer.
    Example: 36 pos → 36*7/50=5.04 → 5,  14 neg → 14*7/50=1.96 → 2,  total=7
    """
    valid   = [s for s in stocks if s.get("contrib") is not None]
    pos     = [s for s in valid if s["contrib"] >= 0]
    neg     = [s for s in valid if s["contrib"] <  0]
    pos_cnt = len(pos)
    neg_cnt = len(neg)
    pos_pts = round(sum(s["contrib"] for s in pos), 2)
    neg_pts = round(sum(s["contrib"] for s in neg), 2)
    pos_sc  = breadth_score(pos_cnt)
    neg_sc  = breadth_score(neg_cnt)
    return {
        "pos_count":       pos_cnt,
        "neg_count":       neg_cnt,
        "pos_contrib_sum": pos_pts,
        "neg_contrib_sum": neg_pts,
        "pos_score":       pos_sc,
        "neg_score":       neg_sc,
        "total_score":     pos_sc + neg_sc,
        "display":         f"{pos_sc} | {neg_sc}",
        "pos_stocks":      sorted(pos, key=lambda x: x["contrib"], reverse=True),
        "neg_stocks":      sorted(neg, key=lambda x: x["contrib"]),
    }

# ════════════════════════════════════════════════════════════
# BUILD SNAPSHOT EXCEL (File 1)
# ════════════════════════════════════════════════════════════
def build_snapshot_excel(label, ist_dt, indices, stocks):
    wb = Workbook(); ws = wb.active
    ws.title = "Snapshot_" + label
    disp = label[:2]+":"+label[2:] if len(label)==4 else label
    time_str = ist_dt.strftime("%d-%b-%Y %H:%M:%S")

    ws.merge_cells("A1:P1")
    c = ws["A1"]; c.value = f"NSE Market Snapshot  —  {disp} IST"
    sc(c, bg=C["title_bg"], fg=C["white"], bold=True, size=14, ha="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:P2")
    ws["A2"].value = f"Captured At (IST):  {time_str}"
    ws["A2"].fill = fill("E3F2FD"); ws["A2"].font = font(size=10, color="333333")
    ws["A2"].alignment = aln("center")

    r = 4
    ws.merge_cells(f"A{r}:D{r}")
    sc(ws.cell(r,1,"📊  INDEX SUMMARY"), bg=C["hdr_dark"], fg=C["white"], bold=True, size=11, ha="center")
    r += 1
    for h, col in zip(["INDEX","LTP (₹)","CHANGE (₹)","% CHANGE"], range(1,5)):
        sc(ws.cell(r,col,h), bg=C["hdr_mid"], fg=C["white"], bold=True, ha="center")

    for i, name in enumerate(["NIFTY 50","SENSEX","BANK NIFTY","NIFTY IT","NIFTY SMALLCAP 50"]):
        r += 1; d = indices.get(name,{}); pct = d.get("pct")
        bg = C["idx_bg"] if i%2==0 else C["idx_alt"]
        pos = pct is not None and pct >= 0
        ws.cell(r,1,name).font = font(bold=True)
        ws.cell(r,1).fill=fill(bg); ws.cell(r,1).border=brd(); ws.cell(r,1).alignment=aln()
        for col, val, fmt in [(2,d.get("ltp"),"#,##0.00"),(3,d.get("chng"),"+#,##0.00;-#,##0.00"),
                               (4,pct and pct/100,"+0.00%;-0.00%")]:
            cell = ws.cell(r,col)
            if pct is not None:
                cell.value=val; cell.number_format=fmt
                cell.font=font(bold=True, color=C["pos_txt"] if pos else C["neg_txt"])
                cell.fill=fill(C["pos_bg"] if pos else C["neg_bg"])
            else:
                cell.value="N/A"; cell.font=font(color="9E9E9E"); cell.fill=fill(bg)
            cell.border=brd(); cell.alignment=aln("center")

    sr = r + 3
    ws.merge_cells(f"A{sr}:G{sr}")
    sc(ws.cell(sr,1,f"📋  ALL NIFTY 50 STOCKS  [{len(stocks)} stocks]"),
       bg=C["hdr_dark"],fg=C["white"],bold=True,size=11,ha="center")
    sr += 1
    for h, col in zip(["#","SYMBOL","LTP (₹)","CHANGE (₹)","% CHANGE","WEIGHT (pts)","CONTRIB (pts)"],range(1,8)):
        sc(ws.cell(sr,col,h),bg="3949AB",fg=C["white"],bold=True,ha="center")

    for i, s in enumerate(stocks):
        sr += 1
        pos = (s["pChng"] or 0) >= 0
        bg  = (C["pos_bg"] if pos else C["neg_bg"]) if i%2==0 else \
              (C["pos_alt"] if pos else C["neg_alt"])
        tc  = C["pos_txt"] if pos else C["neg_txt"]
        ws.cell(sr,1,i+1).fill=fill(bg);ws.cell(sr,1).border=brd();ws.cell(sr,1).alignment=aln("center")
        ws.cell(sr,2,s["symbol"]).font=font(bold=True)
        ws.cell(sr,2).fill=fill(bg);ws.cell(sr,2).border=brd();ws.cell(sr,2).alignment=aln()
        for col,val,fmt in [(3,s["ltp"],"#,##0.00"),(4,s["chng"],"+#,##0.00;-#,##0.00"),
                             (5,s["pChng"] and s["pChng"]/100,"+0.00%;-0.00%"),
                             (6,s.get("weight_pts"),"#,##0.00"),(7,s.get("contrib"),"+0.00;-0.00")]:
            cell=ws.cell(sr,col);cell.value=val
            if val is not None and fmt: cell.number_format=fmt
            cell.font=font(bold=(col in(4,5,7)),color=tc if col in(4,5,7) else "000000")
            cell.fill=fill(bg);cell.border=brd();cell.alignment=aln("center")

    valid=[s for s in stocks if s["pChng"] is not None]
    contrib_valid=[s for s in stocks if s.get("contrib") is not None]
    top7p=sorted(contrib_valid,key=lambda x:x["contrib"],reverse=True)[:7]
    top7n=sorted(contrib_valid,key=lambda x:x["contrib"])[:7]
    trow=r+4
    for top7,col,title,hbg,tc,b1,b2 in [
        (top7p,9,"🟢  TOP 7 POSITIVE",C["grn_hdr"],C["pos_txt"],C["pos_bg"],C["pos_alt"]),
        (top7n,15,"🔴  TOP 7 NEGATIVE",C["red_hdr"],C["neg_txt"],C["neg_bg"],C["neg_alt"]),
    ]:
        ws.merge_cells(start_row=trow,start_column=col,end_row=trow,end_column=col+4)
        sc(ws.cell(trow,col,title),bg=hbg,fg=C["white"],bold=True,size=11,ha="center")
        for h,dc in zip(["SYMBOL","LTP (₹)","CHG (₹)","% CHG","CONTRIB(pts)"],range(col,col+5)):
            sc(ws.cell(trow+1,dc,h),bg=hbg,fg=C["white"],bold=True,ha="center")
        for i,s in enumerate(top7):
            tr=trow+2+i;bg=b1 if i%2==0 else b2
            for dc,val,fmt in [(col,s["symbol"],None),(col+1,s["ltp"],"#,##0.00"),
                                (col+2,s["chng"],"+#,##0.00;-#,##0.00"),
                                (col+3,s["pChng"] and s["pChng"]/100,"+0.00%;-0.00%"),
                                (col+4,s.get("contrib"),"+0.00;-0.00")]:
                cell=ws.cell(tr,dc,val)
                if val is not None and fmt: cell.number_format=fmt
                cell.font=font(bold=True,color=tc);cell.fill=fill(bg)
                cell.border=brd();cell.alignment=aln("center" if dc>col else "left")

    for col,w in {1:4,2:14,3:12,4:12,5:11,6:13,7:13,9:14,10:12,11:11,12:10,13:13,
                  15:14,16:12,17:11,18:10,19:13}.items():
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.freeze_panes="A3"
    return wb

# ════════════════════════════════════════════════════════════
# BUILD LEVELS EXCEL (File 2)
# ════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════
# EMAIL
# ════════════════════════════════════════════════════════════
def send_email(creds, snapshot_path, label, ist_dt, indices, stocks, breadth):
    disp    = label[:2]+":"+label[2:] if len(label)==4 else label
    subject = f"📊 NSE Snapshot {disp} IST — {ist_dt.strftime('%d %b %Y')}"

    valid   = [s for s in stocks if s.get("contrib") is not None]
    top3p   = breadth["pos_stocks"][:3]
    top3n   = breadth["neg_stocks"][:3]

    # Index HTML
    idx_html = ""
    for name in ["NIFTY 50","SENSEX","BANK NIFTY","NIFTY IT","NIFTY SMALLCAP 50"]:
        d=indices.get(name,{}); pct=d.get("pct"); ltp=d.get("ltp")
        if pct is not None:
            color="#1b5e20" if pct>=0 else "#b71c1c"
            bg="#e8f5e9" if pct>=0 else "#ffebee"
            arrow="▲" if pct>=0 else "▼"
            idx_html += (f'<tr style="background:{bg}"><td style="padding:7px 14px;font-weight:bold">{name}</td>'
                f'<td style="padding:7px;color:{color};font-weight:bold;text-align:center">{arrow} {pct:+.2f}%</td>'
                f'<td style="padding:7px;text-align:center">₹{ltp:,.2f}</td></tr>')
        else:
            idx_html += f'<tr><td style="padding:7px 14px;font-weight:bold">{name}</td><td style="padding:7px;color:#9e9e9e;text-align:center">N/A</td><td>—</td></tr>'

    def srows(lst, color, bg):
        return "".join(
            f'<tr style="background:{bg}"><td style="padding:6px 12px;font-weight:bold">{s["symbol"]}</td>'
            f'<td style="padding:6px;color:{color};font-weight:bold;text-align:center">{s["pChng"]:+.2f}%</td>'
            f'<td style="padding:6px;color:{color};text-align:center">{s["contrib"]:+.2f} pts</td></tr>'
            for s in lst)

    note = ('<div style="background:#fff3e0;border-radius:6px;padding:10px 14px;margin-top:12px;'
            'font-size:13px;color:#e65100">⚠️ Market closed or data unavailable.</div>'
            if not valid else "")

    html = f"""<html><body style="font-family:Arial,sans-serif;max-width:640px;margin:auto">
<div style="background:#0d47a1;color:white;padding:18px 22px;border-radius:10px 10px 0 0">
  <h2 style="margin:0;font-size:20px">📊 NSE Snapshot — {disp} IST</h2>
  <p style="margin:5px 0 0;opacity:.85;font-size:13px">{ist_dt.strftime('%d %b %Y  %H:%M:%S IST')}</p>
</div>
<div style="border:1px solid #ddd;border-top:none;padding:18px;border-radius:0 0 10px 10px">

  <!-- BREADTH SCORE BOX -->
  <div style="background:#e8eaf6;border:1px solid #3949ab;border-radius:10px;padding:16px;margin-bottom:16px;text-align:center">
    <div style="font-size:13px;color:#3949ab;font-weight:500;margin-bottom:8px">📊 NIFTY 50 BREADTH SCORE</div>
    <div style="display:flex;justify-content:center;align-items:center;gap:16px">
      <div style="background:#e8f5e9;border:1px solid #2e7d32;border-radius:8px;padding:10px 20px">
        <div style="font-size:11px;color:#2e7d32;font-weight:500">🟢 POSITIVE</div>
        <div style="font-size:28px;font-weight:bold;color:#1b5e20">{breadth["pos_score"]}</div>
        <div style="font-size:11px;color:#555">{breadth["pos_count"]} stocks</div>
        <div style="font-size:12px;color:#1b5e20;font-weight:500">{breadth["pos_contrib_sum"]:+.2f} pts</div>
      </div>
      <div style="font-size:30px;color:#3949ab;font-weight:bold">|</div>
      <div style="background:#ffebee;border:1px solid #c62828;border-radius:8px;padding:10px 20px">
        <div style="font-size:11px;color:#c62828;font-weight:500">🔴 NEGATIVE</div>
        <div style="font-size:28px;font-weight:bold;color:#b71c1c">{breadth["neg_score"]}</div>
        <div style="font-size:11px;color:#555">{breadth["neg_count"]} stocks</div>
        <div style="font-size:12px;color:#b71c1c;font-weight:500">{breadth["neg_contrib_sum"]:+.2f} pts</div>
      </div>
      <div style="font-size:30px;color:#3949ab;font-weight:bold">=</div>
      <div style="background:#e8eaf6;border:1px solid #3949ab;border-radius:8px;padding:10px 20px">
        <div style="font-size:11px;color:#3949ab;font-weight:500">TOTAL</div>
        <div style="font-size:28px;font-weight:bold;color:#0d47a1">{breadth["total_score"]}</div>
        <div style="font-size:11px;color:#555">always 7</div>
      </div>
    </div>
  </div>

  <h3 style="color:#0d47a1;margin:0 0 8px">Index Performance</h3>
  <table width="100%" cellspacing="0" style="border-collapse:collapse;border:1px solid #e0e0e0">
    <tr style="background:#1976d2;color:white">
      <th style="padding:8px 14px;text-align:left">Index</th>
      <th style="padding:8px">% Change</th><th style="padding:8px">LTP</th>
    </tr>{idx_html}
  </table>

  <table width="100%" style="margin-top:16px;border-collapse:collapse"><tr valign="top">
    <td width="50%" style="padding-right:8px">
      <h3 style="color:#2e7d32;margin:0 0 8px">🏆 Top 3 Boosters</h3>
      <table width="100%" cellspacing="0" style="border-collapse:collapse;border:1px solid #e0e0e0">
        {srows(top3p,"#1b5e20","#e8f5e9") if top3p else '<tr><td style="padding:8px;color:#9e9e9e">No data</td></tr>'}
      </table>
    </td>
    <td width="50%" style="padding-left:8px">
      <h3 style="color:#c62828;margin:0 0 8px">📉 Top 3 Draggers</h3>
      <table width="100%" cellspacing="0" style="border-collapse:collapse;border:1px solid #e0e0e0">
        {srows(top3n,"#b71c1c","#ffebee") if top3n else '<tr><td style="padding:8px;color:#9e9e9e">No data</td></tr>'}
      </table>
    </td>
  </tr></table>
  {note}
  <div style="margin-top:16px;padding:12px 14px;background:#f5f5f5;border-radius:8px;font-size:13px">
    📎 Excel with all 50 stocks attached<br>
    ☁️ Saved to Google Drive → NSE Snapshots folder
  </div>
</div></body></html>"""

    msg = MIMEMultipart("alternative")
    msg["From"]=creds["gmail_sender"]; msg["To"]=creds["recipient"]; msg["Subject"]=subject
    msg.attach(MIMEText(html,"html"))
    with open(snapshot_path,"rb") as f:
        part=MIMEBase("application","vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read()); encoders.encode_base64(part)
        part.add_header("Content-Disposition","attachment",filename=os.path.basename(snapshot_path))
        msg.attach(part)
    with smtplib.SMTP_SSL("smtp.gmail.com",465) as s:
        s.login(creds["gmail_sender"],creds["gmail_pass"])
        s.sendmail(creds["gmail_sender"],creds["recipient"],msg.as_string())
    print(f"✅ Email sent to {creds['recipient']}")


def upload_drive(creds, xlsx_path):
    folder_id = creds["drive_folder"]
    if not folder_id or len(folder_id) < 10:
        print("❌ Drive folder ID missing"); return
    try:
        svc = build("drive","v3",credentials=get_sa_creds(
            creds["sa_json"],["https://www.googleapis.com/auth/drive"]))
        name = os.path.basename(xlsx_path)
        q    = f"name='{name}' and '{folder_id}' in parents and trashed=false"
        for f in svc.files().list(q=q,fields="files(id)",supportsAllDrives=True,
                includeItemsFromAllDrives=True).execute().get("files",[]):
            svc.files().delete(fileId=f["id"],supportsAllDrives=True).execute()
        meta  = {"name":name,"parents":[folder_id]}
        media = MediaFileUpload(xlsx_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=False)
        res = svc.files().create(body=meta,media_body=media,
            fields="id,webViewLink",supportsAllDrives=True).execute()
        print(f"✅ Drive: {name}  →  {res.get('webViewLink','')}")
    except Exception as e:
        print(f"❌ Drive upload failed: {e}")

# ════════════════════════════════════════════════════════════
# GOOGLE SHEETS
# ════════════════════════════════════════════════════════════
def update_sheets(creds, label, ist_dt, indices, stocks, breadth):
    try:
        gc  = gspread.authorize(get_sa_creds(
            creds["sa_json"],["https://www.googleapis.com/auth/spreadsheets"]))
        sh  = gc.open_by_key(creds["sheet_id"])
        disp = label[:2]+":"+label[2:] if len(label)==4 else label

        tab1 = "Snapshot_" + label
        try: sh.del_worksheet(sh.worksheet(tab1))
        except: pass
        ws1 = sh.add_worksheet(title=tab1, rows=120, cols=20)

        rows1 = [
            [f"NSE Snapshot — {disp} IST"]+[""]*19,
            [f"Captured: {ist_dt.strftime('%d-%b-%Y %H:%M:%S IST')}"]+[""]*19,
            [""]*20,
            ["BREADTH SCORE","POSITIVE","NEGATIVE","TOTAL",""]+[""]*15,
            [f"Score",breadth["pos_score"],breadth["neg_score"],breadth["total_score"],"",
             f"Stocks",breadth["pos_count"],breadth["neg_count"],breadth["pos_count"]+breadth["neg_count"],"",
             f"Contrib pts",breadth["pos_contrib_sum"],breadth["neg_contrib_sum"],
             round(breadth["pos_contrib_sum"]+breadth["neg_contrib_sum"],2)]+[""]*5,
            [""]*20,
            ["INDEX","LTP","CHG","% CHG",""]+[""]*15,
        ]
        for name in ["NIFTY 50","SENSEX","BANK NIFTY","NIFTY IT","NIFTY SMALLCAP 50"]:
            d=indices.get(name,{}); p=d.get("pct")
            rows1.append([name,d.get("ltp","N/A"),d.get("chng","N/A"),
                          f"{p:+.2f}%" if p is not None else "N/A",""]+[""]*15)
        rows1.append([""]*20)
        rows1.append(["#","SYMBOL","LTP","CHG","% CHG","WT(pts)","CONTRIB","",
                      "TOP POSITIVE","LTP","CHG","% CHG","CONTRIB","",
                      "TOP NEGATIVE","LTP","CHG","% CHG","CONTRIB"])

        top7p = breadth["pos_stocks"][:7]
        top7n = breadth["neg_stocks"][:7]
        all_stocks = list(enumerate(stocks))
        for i, s in all_stocks:
            p=s["pChng"]
            row=[i+1,s["symbol"],s.get("ltp",""),
                 f'{s["chng"]:+.2f}' if s["chng"] is not None else "",
                 f'{p:+.2f}%' if p is not None else "",
                 f'{s["weight_pts"]:.2f}' if s.get("weight_pts") is not None else "",
                 f'{s["contrib"]:+.2f}' if s.get("contrib") is not None else "",""]
            row+=([top7p[i]["symbol"],top7p[i]["ltp"],
                   f'{top7p[i]["chng"]:+.2f}',f'{top7p[i]["pChng"]:+.2f}%',
                   f'{top7p[i]["contrib"]:+.2f}',""] if i<len(top7p) else [""]*6)
            row+=([top7n[i]["symbol"],top7n[i]["ltp"],
                   f'{top7n[i]["chng"]:+.2f}',f'{top7n[i]["pChng"]:+.2f}%',
                   f'{top7n[i]["contrib"]:+.2f}'] if i<len(top7n) else [""]*5)
            rows1.append(row)
        ws1.update("A1", rows1)
        print(f"✅ Sheets updated: {tab1}")
    except Exception as e:
        print(f"⚠️  Sheets failed: {e}")
        traceback.print_exc()


# ════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════
def main():
    ist_dt    = datetime.now(IST)
    label     = env("MANUAL_LABEL") or ist_dt.strftime("%H%M")
    disp      = label[:2]+":"+label[2:] if len(label)==4 else label

    print(f"\n{'='*60}")
    print(f"  NSE Snapshot + Levels — {disp} IST | {ist_dt.strftime('%d-%b-%Y')}")
    print(f"{'='*60}")

    creds      = get_creds()
    obj        = angel_login(creds)
    token_map  = load_token_map()
    indices    = fetch_indices(obj)

    nifty_ltp   = indices.get("NIFTY 50",{}).get("ltp")   or 0
    prev_close  = indices.get("NIFTY 50",{}).get("close") or 0
    print(f"\n  Nifty LTP: {nifty_ltp}  |  Prev Close: {prev_close}")

    stocks = fetch_nifty50(obj, nifty_ltp, token_map)

    # Calculate breadth score
    print("\n📊 Calculating breadth score...")
    breadth = calc_breadth(stocks)
    print(f"  Positive stocks : {breadth['pos_count']}  →  score {breadth['pos_score']}")
    print(f"  Negative stocks : {breadth['neg_count']}  →  score {breadth['neg_score']}")
    print(f"  Total score     : {breadth['total_score']} (always 7)")
    print(f"  +ve contrib sum : {breadth['pos_contrib_sum']:+.2f} pts")
    print(f"  -ve contrib sum : {breadth['neg_contrib_sum']:+.2f} pts")

    os.makedirs("output", exist_ok=True)
    date_str = ist_dt.strftime("%Y-%m-%d")

    # Build Excel 1 — Snapshot
    print("\n📁 Building Snapshot Excel...")
    wb1 = build_snapshot_excel(label, ist_dt, indices, stocks)
    snap_path = f"output/NSE_{date_str}_{label}.xlsx"
    wb1.save(snap_path)
    print(f"  ✅ {snap_path}")

    # Send email
    print("\n📧 Sending email...")
    send_email(creds, snap_path, label, ist_dt, indices, stocks, breadth)

    # Upload to Drive
    print("\n☁️  Uploading to Drive...")
    upload_drive(creds, snap_path)

    # Update Sheets
    print("\n📊 Updating Google Sheets...")
    update_sheets(creds, label, ist_dt, indices, stocks, breadth)

    print(f"\n✅ ALL DONE — {disp} IST\n")
